"""
Excel parser using openpyxl with trio support.
"""
import logging
from typing import List, Tuple, Optional, Dict, Any
from pathlib import Path

import trio
import pandas as pd
from openpyxl import load_workbook

from .base_parser import BaseParser, DocumentStructure, ParsingError

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """
    Excel parser using openpyxl and pandas.
    """
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
        self.parser_name = "ExcelParser"
    
    async def parse(self, file_path: str, **kwargs) -> DocumentStructure:
        """
        Parse Excel document.
        
        Args:
            file_path: Path to Excel file
            **kwargs: Additional options:
                - sheet_names: List of sheet names to parse (None for all)
                - max_rows: Maximum rows to process per sheet
                - detect_tables: bool
                
        Returns:
            DocumentStructure with parsed content
        """
        try:
            # Validate file
            if not await self.validate_file(file_path):
                raise ParsingError(f"Invalid Excel file: {file_path}", self.parser_name, file_path)
            
            # Parse options
            sheet_names = kwargs.get('sheet_names', None)
            max_rows = kwargs.get('max_rows', 10000)
            detect_tables = kwargs.get('detect_tables', True)
            
            logger.info(f"Parsing Excel {file_path}")
            
            # Parse Excel (run in thread)
            doc_structure = await trio.to_thread.run_sync(
                self._parse_excel_sync,
                file_path,
                sheet_names,
                max_rows,
                detect_tables
            )
            
            logger.info(f"Successfully parsed Excel {file_path}: {len(doc_structure.text_blocks)} text blocks")
            
            return doc_structure
            
        except Exception as e:
            logger.error(f"Failed to parse Excel {file_path}: {e}")
            raise ParsingError(f"Excel parsing failed: {str(e)}", self.parser_name, file_path)
    
    def _parse_excel_sync(self, file_path: str, sheet_names: Optional[List[str]], 
                         max_rows: int, detect_tables: bool) -> DocumentStructure:
        """
        Synchronous Excel parsing.
        """
        doc_structure = DocumentStructure()
        doc_structure.metadata = self._extract_metadata(file_path)
        
        try:
            # Load workbook
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheets to process
            sheets_to_process = sheet_names if sheet_names else workbook.sheetnames
            
            for sheet_name in sheets_to_process:
                if sheet_name not in workbook.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found in {file_path}")
                    continue
                
                worksheet = workbook[sheet_name]
                
                # Read data using pandas for better handling
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows)
                
                if df.empty:
                    continue
                
                # Process sheet data
                sheet_data = self._process_sheet_data(df, sheet_name, detect_tables)
                
                # Add to document structure
                doc_structure.text_blocks.extend(sheet_data["text_blocks"])
                doc_structure.tables.extend(sheet_data["tables"])
                
                # Add sheet as page
                doc_structure.pages.append({
                    "page_number": len(doc_structure.pages) + 1,
                    "sheet_name": sheet_name,
                    "content": sheet_data["content"],
                    "rows": len(df),
                    "columns": len(df.columns)
                })
            
            workbook.close()
            
        except Exception as e:
            raise ParsingError(f"Failed to load Excel file: {str(e)}", self.parser_name, file_path)
        
        # Update metadata
        doc_structure.metadata.update({
            "total_sheets": len(doc_structure.pages),
            "total_tables": len(doc_structure.tables),
            "total_text_blocks": len(doc_structure.text_blocks)
        })
        
        return doc_structure
    
    def _process_sheet_data(self, df: pd.DataFrame, sheet_name: str, detect_tables: bool) -> Dict[str, Any]:
        """
        Process sheet data into text blocks and tables.
        """
        text_blocks = []
        tables = []
        content_parts = []
        
        # Clean DataFrame
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        if df.empty:
            return {
                "text_blocks": text_blocks,
                "tables": tables,
                "content": ""
            }
        
        if detect_tables:
            # Treat entire sheet as a table if it has structure
            if len(df.columns) > 1 and len(df) > 1:
                table_content = self._dataframe_to_text(df)
                
                table_data = {
                    "content": table_content,
                    "sheet_name": sheet_name,
                    "rows": len(df),
                    "columns": len(df.columns),
                    "headers": df.columns.tolist(),
                    "element_type": "table"
                }
                
                tables.append(table_data)
                
                # Add as text block
                metadata = {
                    "element_type": "table",
                    "sheet_name": sheet_name,
                    "rows": len(df),
                    "columns": len(df.columns)
                }
                text_blocks.append((table_content, metadata))
                content_parts.append(table_content)
            
            else:
                # Process as individual cells
                for idx, row in df.iterrows():
                    for col, value in row.items():
                        if pd.notna(value) and str(value).strip():
                            cell_text = str(value).strip()
                            metadata = {
                                "element_type": "cell",
                                "sheet_name": sheet_name,
                                "row": idx,
                                "column": col
                            }
                            text_blocks.append((cell_text, metadata))
                            content_parts.append(cell_text)
        
        else:
            # Process all values as text
            for idx, row in df.iterrows():
                row_text = " | ".join([str(val) for val in row.values if pd.notna(val)])
                if row_text.strip():
                    metadata = {
                        "element_type": "row",
                        "sheet_name": sheet_name,
                        "row": idx
                    }
                    text_blocks.append((row_text, metadata))
                    content_parts.append(row_text)
        
        return {
            "text_blocks": text_blocks,
            "tables": tables,
            "content": "\n".join(content_parts)
        }
    
    def _dataframe_to_text(self, df: pd.DataFrame) -> str:
        """
        Convert DataFrame to readable text format.
        """
        # Create header
        headers = " | ".join(str(col) for col in df.columns)
        separator = "-" * len(headers)
        
        # Create rows
        rows = []
        for _, row in df.iterrows():
            row_text = " | ".join(str(val) if pd.notna(val) else "" for val in row.values)
            rows.append(row_text)
        
        return "\n".join([headers, separator] + rows)